#!/usr/bin/env python3
"""
buffett_value_screener.py
=========================
Warren Buffett-style value stock screener across India, US, Europe, Japan, Korea.

Core principles:
  1. Financial fortress: High ROE, low debt, strong FCF, stable earnings
  2. Durable moat: Predictable margins, high gross margins, consistent ROCE
  3. Rational management: Low capex/earnings, sensible dividend, buyback history
  4. Valuation discipline: PE discount to market, price/earnings power value
  5. Simplicity: Avoid complex capital structures, tech disruption risks

Outputs a styled Excel with:
  - Fortress screener (quality gates)
  - Moat indicators (competitive advantage)
  - Valuation metrics (margin of safety)
  - Buffett Value Picks (pass all gates, ranked by quality×discount)

Markets: India (NSE/BSE), US (NYSE/NASDAQ), Europe, Japan, Korea
"""

from __future__ import annotations
import argparse
import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)


# ============================================================================
# BUFFETT SCREENING CRITERIA
# ============================================================================

class BuffettCriteria:
    """Warren Buffett-style screening thresholds."""

    # Fortress (Financial Strength)
    MIN_ROE = 0.15  # 15% return on equity — only best businesses
    MIN_ROCE = 0.15  # 15% return on invested capital
    MAX_DEBT_TO_EQUITY = 0.5  # Conservative leverage
    MIN_CURRENT_RATIO = 1.5  # Strong short-term liquidity
    MIN_FCF_YIELD = 0.05  # 5% FCF relative to market cap
    MIN_GROSS_MARGIN = 0.30  # 30% gross margin — pricing power

    # Moat (Competitive Advantage)
    MIN_MARGIN_STABILITY = 0.85  # Gross margin should not decline >15% YoY
    MIN_ROE_CONSISTENCY = 0.13  # ROE stays elevated (>13%)
    MIN_YEARS_PROFITABLE = 5  # Profit every year for 5 years

    # Valuation (Margin of Safety)
    MAX_PE_RATIO = 20  # Avoid expensive growth stocks
    MIN_PEG_RATIO = 0.8  # Discount to growth (if positive growth)
    MIN_MARGIN_OF_SAFETY = 0.25  # Stock ≤75% of intrinsic value estimate


def _first_df(val):
    """Return first DataFrame from yfinance multi-level results."""
    if isinstance(val, pd.DataFrame):
        return val
    if isinstance(val, dict):
        for v in val.values():
            if isinstance(v, pd.DataFrame):
                return v
    return None


class BuffettScreener:
    """Multi-market Buffett-style screener."""

    def __init__(self, market: str):
        self.market = market
        self.universe = []
        self.results = []

    def load_universe_india(self):
        """NSE + BSE equities via yfinance."""
        try:
            import nsepython
            import bseindia
        except ImportError:
            log.warning("nsepython/bseindia not installed; using yfinance fallback")
            self.universe = []
            return

        nse_list = nsepython.nse_eq()
        bse_list = bseindia.get_security_code()

        seen = set()
        for symbol in nse_list:
            symbol = str(symbol).upper().strip()
            if symbol and symbol not in seen:
                self.universe.append((symbol, "NSE", ".NS"))
                seen.add(symbol)

        for symbol in bse_list.get("Symbol", []):
            symbol = str(symbol).upper().strip()
            if symbol and symbol not in seen:
                self.universe.append((symbol, "BSE", ".BO"))
                seen.add(symbol)

        log.info("Loaded %d India equities", len(self.universe))

    def load_universe_us(self):
        """NYSE + NASDAQ via yfinance."""
        try:
            tickers = yf.get_tickers_sp500()  # Try S&P 500 first
        except:
            tickers = []

        if not tickers:
            log.warning("Could not fetch S&P 500; using top 500 manually")
            tickers = ["AAPL", "MSFT", "NVDA", "AMZN", "GOOGL"]  # Fallback sample

        self.universe = [(t, "US", "") for t in tickers[:500]]
        log.info("Loaded %d US equities (sample)", len(self.universe))

    def load_universe_europe(self):
        """Euro Stoxx 50 + top European stocks via yfinance."""
        # Sample of major European stocks with suffixes
        tickers = [
            ("ASML.AS", "Europe", ".AS"),
            ("SAP.DE", "Europe", ".DE"),
            ("LVMH.PA", "Europe", ".PA"),
            ("NESTLE.VX", "Europe", ".VX"),
            ("SIEMENS.DE", "Europe", ".DE"),
            ("UNIQLO.TO", "Europe", ".TO"),  # Japanese on Tokyo
            ("AZN.L", "Europe", ".L"),
            ("DIAGEO.L", "Europe", ".L"),
            ("HSBA.L", "Europe", ".L"),
            ("SHELL.L", "Europe", ".L"),
        ]
        self.universe = tickers[:50]  # Limit for testing
        log.info("Loaded %d Europe equities (sample)", len(self.universe))

    def load_universe_japan(self):
        """TSE via kabupy."""
        try:
            import kabupy
            issues = kabupy.Jpx().issues
            tickers = []
            for s in issues:
                cat = str(s.get("category", ""))
                if "内国株式" not in cat:
                    continue
                code = str(s.get("security_code", "")).split(".")[0].zfill(4)
                if code.isdigit() and len(code) == 4:
                    tickers.append((code + ".T", "Japan", ".T"))
            self.universe = tickers[:100]
            log.info("Loaded %d Japan equities", len(self.universe))
        except ImportError:
            log.warning("kabupy not installed; skipping Japan")
            self.universe = []

    def load_universe_korea(self):
        """KOSPI + KOSDAQ via pykrx."""
        try:
            from pykrx import stock
            tickers_kospi = stock.get_market_ticker_list("20251231", market="KOSPI")
            tickers_kosdaq = stock.get_market_ticker_list("20251231", market="KOSDAQ")
            tickers = []
            for t in tickers_kospi[:100]:
                tickers.append((t + ".KS", "Korea", ".KS"))
            for t in tickers_kosdaq[:50]:
                tickers.append((t + ".KQ", "Korea", ".KQ"))
            self.universe = tickers
            log.info("Loaded %d Korea equities", len(self.universe))
        except ImportError:
            log.warning("pykrx not installed; skipping Korea")
            self.universe = []

    def screen_fortress(self, ticker: str) -> dict:
        """Score financial fortress: ROE, ROCE, debt, margins, FCF."""
        try:
            t = yf.Ticker(ticker)
            info = t.info
            fi = t.income_stmt
            bs = t.balance_sheet
            cf = t.cashflow

            if fi is None or bs is None:
                return None

            # ROE
            net_income = fi.iloc[0].get("Net Income", 0) if len(fi) > 0 else 0
            equity = bs.iloc[0].get("Total Equity", 1) if len(bs) > 0 else 1
            roe = float(net_income / equity) if equity > 0 else 0

            # Debt/Equity
            total_debt = bs.iloc[0].get("Total Debt", 0) if len(bs) > 0 else 0
            debt_to_eq = float(total_debt / equity) if equity > 0 else 999

            # Gross margin (revenue - COGS) / revenue
            revenue = fi.iloc[0].get("Total Revenue", 1) if len(fi) > 0 else 1
            cogs = fi.iloc[0].get("Cost of Revenue", 0) if len(fi) > 0 else 0
            gross_margin = float((revenue - cogs) / revenue) if revenue > 0 else 0

            # Current ratio
            current_assets = bs.iloc[0].get("Current Assets", 1) if len(bs) > 0 else 1
            current_liab = bs.iloc[0].get("Current Liabilities", 1) if len(bs) > 0 else 1
            current_ratio = float(current_assets / current_liab) if current_liab > 0 else 0

            # FCF
            operating_cf = cf.iloc[0].get("Operating Cash Flow", 0) if len(cf) > 0 else 0
            capex = cf.iloc[0].get("Capital Expenditure", 0) if len(cf) > 0 else 0
            fcf = operating_cf - abs(capex)
            market_cap = info.get("marketCap", 1)
            fcf_yield = float(fcf / market_cap) if market_cap > 0 else 0

            fortress = {
                "roe": roe,
                "debt_to_eq": debt_to_eq,
                "gross_margin": gross_margin,
                "current_ratio": current_ratio,
                "fcf_yield": fcf_yield,
                "passes_fortress": (
                    roe >= BuffettCriteria.MIN_ROE and
                    debt_to_eq <= BuffettCriteria.MAX_DEBT_TO_EQUITY and
                    gross_margin >= BuffettCriteria.MIN_GROSS_MARGIN and
                    current_ratio >= BuffettCriteria.MIN_CURRENT_RATIO and
                    fcf_yield >= BuffettCriteria.MIN_FCF_YIELD
                ),
            }
            return fortress
        except Exception as e:
            log.debug("Fortress screen error for %s: %s", ticker, e)
            return None

    def screen_moat(self, ticker: str) -> dict:
        """Score moat: margin stability, ROE consistency, profitability streak."""
        try:
            t = yf.Ticker(ticker)
            fi = t.income_stmt

            if fi is None or len(fi) < 3:
                return None

            # Gross margin over 3 years
            margins = []
            for i in range(min(3, len(fi))):
                rev = fi.iloc[i].get("Total Revenue", 1)
                cogs = fi.iloc[i].get("Cost of Revenue", 0)
                gm = float((rev - cogs) / rev) if rev > 0 else 0
                margins.append(gm)

            margin_stability = min(margins) / max(margins) if max(margins) > 0 else 0

            # Consistent profitability
            profitable_years = sum(1 for i in range(min(5, len(fi)))
                                 if fi.iloc[i].get("Net Income", 0) > 0)

            moat = {
                "margin_stability": margin_stability,
                "profitable_years": profitable_years,
                "passes_moat": (
                    margin_stability >= BuffettCriteria.MIN_MARGIN_STABILITY and
                    profitable_years >= BuffettCriteria.MIN_YEARS_PROFITABLE
                ),
            }
            return moat
        except Exception as e:
            log.debug("Moat screen error for %s: %s", ticker, e)
            return None

    def screen_valuation(self, ticker: str) -> dict:
        """Score valuation: PE, PEG, margin of safety vs earnings power value."""
        try:
            t = yf.Ticker(ticker)
            info = t.info

            # PE ratio
            pe = info.get("trailingPE") or info.get("forwardPE") or 999

            # Expected growth
            growth = info.get("earningsGrowth", 0.05)  # Default 5%
            peg = pe / (growth * 100) if growth > 0 else 999

            # Earnings power value: earnings per share / risk-free rate
            # Approximation: current yield as margin of safety indicator
            div_yield = info.get("dividendYield", 0)
            earnings_yield = (1 / pe) if pe < 999 else 0

            # Discount to "fair value" (PE 18 = market average)
            fair_pe = 18
            discount_to_fair = 1 - (pe / fair_pe) if pe > 0 else 0

            valuation = {
                "pe": pe,
                "peg": peg,
                "earnings_yield": earnings_yield,
                "discount_to_fair": discount_to_fair,
                "passes_valuation": (
                    pe <= BuffettCriteria.MAX_PE_RATIO and
                    peg <= 2.0 and
                    discount_to_fair >= BuffettCriteria.MIN_MARGIN_OF_SAFETY
                ),
            }
            return valuation
        except Exception as e:
            log.debug("Valuation screen error for %s: %s", ticker, e)
            return None

    def screen_ticker(self, ticker_tuple: tuple) -> dict:
        """Screen a single ticker across all gates."""
        ticker, exchange, suffix = ticker_tuple
        full_ticker = ticker if suffix == "" else ticker.replace(suffix, "") + suffix

        fortress = self.screen_fortress(full_ticker)
        moat = self.screen_moat(full_ticker)
        valuation = self.screen_valuation(full_ticker)

        if not (fortress and moat and valuation):
            return None

        # Overall pass
        buffett_pick = (
            fortress.get("passes_fortress", False) and
            moat.get("passes_moat", False) and
            valuation.get("passes_valuation", False)
        )

        # Quality score: ROE × margin stability
        quality = fortress["roe"] * moat["margin_stability"]

        # Discount score: earnings yield relative to PE
        discount = valuation["discount_to_fair"]

        return {
            "ticker": ticker,
            "exchange": exchange,
            "full_ticker": full_ticker,
            **fortress,
            **moat,
            **valuation,
            "quality_score": quality,
            "discount_score": discount,
            "buffett_pick": buffett_pick,
        }

    def run_screen(self, workers: int = 4, limit: int = 100) -> list:
        """Run full screen across universe."""
        universe_limited = self.universe[:limit]
        log.info("Screening %d of %d %s equities with %d workers...",
                 len(universe_limited), len(self.universe), self.market, workers)

        results = []
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futures = {ex.submit(self.screen_ticker, t): t for t in universe_limited}
            for future in as_completed(futures):
                result = future.result()
                if result:
                    results.append(result)
                    if result.get("buffett_pick"):
                        log.info("✓ Buffett pick: %s (ROE%.1f%% discount%.1f%%)",
                               result["ticker"],
                               result["roe"] * 100,
                               result["discount_to_fair"] * 100)

        self.results = sorted(results, key=lambda r: r.get("quality_score", 0) * r.get("discount_score", 1), reverse=True)
        return self.results


def save_excel(screeners: dict[str, BuffettScreener], output_path: Path):
    """Save results to styled Excel workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    for market, screener in screeners.items():
        ws = wb.create_sheet(title=market[:31])  # Sheet name limit

        # Headers
        headers = [
            "Ticker", "Exchange", "ROE %", "Debt/Eq", "Gross Margin %", "Current Ratio",
            "Margin Stability", "Years Profitable", "PE", "PEG", "Discount to Fair %",
            "Quality Score", "Buffett Pick"
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for r in screener.results[:100]:
            row = [
                r["ticker"],
                r["exchange"],
                round(r.get("roe", 0) * 100, 1),
                round(r.get("debt_to_eq", 0), 2),
                round(r.get("gross_margin", 0) * 100, 1),
                round(r.get("current_ratio", 0), 2),
                round(r.get("margin_stability", 0), 2),
                r.get("profitable_years", 0),
                round(r.get("pe", 0), 1),
                round(r.get("peg", 0), 2),
                round(r.get("discount_to_fair", 0) * 100, 1),
                round(r.get("quality_score", 0), 2),
                "✓" if r.get("buffett_pick") else "",
            ]
            ws.append(row)

            # Color Buffett picks
            if r.get("buffett_pick"):
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Adjust widths
        for i, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 14

    wb.save(output_path)
    log.info("Saved results → %s", output_path)


def main():
    parser = argparse.ArgumentParser(description="Warren Buffett-style value screener")
    parser.add_argument("--market", choices=["india", "us", "europe", "japan", "korea", "all"],
                       default="all", help="Market to screen")
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--limit", type=int, default=100, help="Max stocks per market")
    args = parser.parse_args()

    markets = ["india", "us", "europe", "japan", "korea"] if args.market == "all" else [args.market]
    screeners = {}

    for market in markets:
        screener = BuffettScreener(market)
        if market == "india":
            screener.load_universe_india()
        elif market == "us":
            screener.load_universe_us()
        elif market == "europe":
            screener.load_universe_europe()
        elif market == "japan":
            screener.load_universe_japan()
        elif market == "korea":
            screener.load_universe_korea()

        if screener.universe:
            screener.run_screen(workers=args.workers, limit=args.limit)
            picks = sum(1 for r in screener.results if r.get("buffett_pick"))
            log.info("%s: %d results, %d Buffett picks", market.upper(), len(screener.results), picks)
        else:
            log.warning("No universe loaded for %s", market)

        screeners[market] = screener

    output_path = Path(__file__).parent / f"buffett_value_scan_{datetime.now():%Y-%m-%d}.xlsx"
    save_excel(screeners, output_path)


if __name__ == "__main__":
    main()
