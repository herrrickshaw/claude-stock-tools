# europe_market_scan.py
# =====================
# Full European equity universe scanner — 1,700+ stocks across 17 exchanges.
#
# Exchanges covered:
#   Euronext (Paris .PA, Amsterdam .AS, Brussels .BR, Lisbon .LS,
#             Milan .MI, Dublin .IR, Oslo .OL)
#   Deutsche Boerse (Xetra .DE, Frankfurt .F)
#   Nasdaq Nordic & Baltic (Stockholm .ST, Helsinki .HE, Copenhagen .CO,
#                           Vilnius .VS, Tallinn .TL, Riga .RG)
#   BME Madrid (.MC), Warsaw (.WA), Athens (.AT), Vienna (.VI),
#   LSE (.L), Prague (.PR), Bucharest (.RO), Budapest (.BD), SIX (.SW)
#
# Universe source:
#   EU_All_Listed_Companies_Full_Jun2026.xlsx  (1,851 rows, 147 US-listed excluded)
#   Falls back to a live fetch from companiesmarketcap.com if the file is absent.
#
# Pipeline:
#   Stage 1 — Load EU equity universe (1,704 tickers with name / country / sector)
#   Stage 2 — Bulk OHLC download via yfinance (batched, 200 tickers per call)
#   Stage 3 — Darvas Box screen + 200-day MA trend on every stock
#   Stage 4 — Piotroski F-Score + Coffee Can on Darvas BREAKOUT candidates only
#   Stage 5 — Save styled 4-sheet Excel workbook
#
# Output sheets:
#   All_Stocks      — price + Darvas signal for every stock scanned
#   Darvas_Signals  — breakout / breakdown alerts ranked by upside
#   Fundamentals    — Piotroski + Coffee Can results for breakout candidates
#   Triple_Hits     — BREAKOUT_BUY + Piotroski >= 7 + Coffee Can PASS
#
# Usage:
#   python europe_market_scan.py
#   python europe_market_scan.py --top 200            # limit to first 200 tickers
#   python europe_market_scan.py --min-cap 1          # only mcap >= $1B
#   python europe_market_scan.py --exchange PA DE ST  # filter by suffix
#   python europe_market_scan.py --no-scans           # Darvas only
#   python europe_market_scan.py --workers 10
#
# Install:
#   pip install yfinance pandas openpyxl requests

import argparse
import json
import sys
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
import yfinance as yf

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Configuration ─────────────────────────────────────────────────────────────
DOWNLOAD_DIR     = Path("./europe_scan")
DOWNLOAD_DIR.mkdir(exist_ok=True)

DARVAS_CONFIRM      = 3
BATCH_SIZE          = 200
SLEEP_BETWEEN       = 1.5
MAX_WORKERS         = 10
PIOTROSKI_STRONG    = 7
MAX_FUND_CANDIDATES = 300
SYMBOL_CACHE_TTL    = 86400   # 24 h

_CACHE_FILE = DOWNLOAD_DIR / ".symbols_cache.json"

# Default path to the pre-built universe file
_DEFAULT_UNIVERSE_FILE = (
    Path(__file__).resolve().parent.parent.parent
    / "Downloads"
    / "EU_All_Listed_Companies_Full_Jun2026.xlsx"
)

EXCHANGE_NAMES = {
    ".PA": "Euronext Paris",       ".AS": "Euronext Amsterdam",
    ".BR": "Euronext Brussels",    ".LS": "Euronext Lisbon",
    ".MI": "Euronext Milan",       ".IR": "Euronext Dublin",
    ".OL": "Euronext Oslo",        ".DE": "Xetra",
    ".F":  "Deutsche Boerse",      ".MC": "BME Madrid",
    ".ST": "Nasdaq Stockholm",     ".HE": "Nasdaq Helsinki",
    ".CO": "Nasdaq Copenhagen",    ".VS": "Nasdaq Vilnius",
    ".TL": "Nasdaq Tallinn",       ".RG": "Nasdaq Riga",
    ".WA": "Warsaw",               ".AT": "Athens",
    ".VI": "Vienna",               ".L":  "LSE",
    ".SW": "SIX Swiss",            ".PR": "Prague",
    ".RO": "Bucharest",            ".BD": "Budapest",
}

# ── Universe ──────────────────────────────────────────────────────────────────

def _load_cache():
    try:
        if _CACHE_FILE.exists():
            data = json.loads(_CACHE_FILE.read_text())
            if time.time() - data.get("ts", 0) < SYMBOL_CACHE_TTL:
                return data.get("universe", [])
    except Exception:
        pass
    return None


def _save_cache(universe):
    try:
        _CACHE_FILE.write_text(json.dumps({"ts": time.time(), "universe": universe}))
    except Exception:
        pass


def _ticker_suffix(ticker: str) -> str:
    dot = ticker.rfind(".")
    return ("." + ticker[dot + 1:]) if dot >= 0 else ""


def load_universe_from_file(path: Path) -> list[dict]:
    """Parse EU universe from the pre-built Excel file."""
    db = pd.read_excel(path, sheet_name="Database")
    eu = db[db["Exchange Group"] != "US-listed (EU-domiciled)"].copy()
    eu = eu.dropna(subset=["Ticker"])
    result = []
    for _, row in eu.iterrows():
        ticker = str(row["Ticker"]).strip()
        if not ticker or ticker == "nan":
            continue
        result.append({
            "ticker":   ticker,
            "name":     str(row.get("Company", "")).strip(),
            "country":  str(row.get("Country", "")).strip(),
            "sector":   str(row.get("GICS Sector (top 200)", "—")).strip(),
            "exchange": str(row.get("Exchange", "")).strip(),
            "mcap_b":   float(row["Market Cap ($B)"]) if pd.notna(row.get("Market Cap ($B)")) else None,
            "cap_band": str(row.get("Cap Band", "")).strip(),
        })
    return result


def fetch_universe_live() -> list[dict]:
    """Fallback: scrape companiesmarketcap.com CSV for top EU stocks."""
    print("  Fetching live EU universe from companiesmarketcap.com …")
    url = "https://companiesmarketcap.com/europe/largest-companies-in-europe-by-market-cap/?download=csv"
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
        r.raise_for_status()
        df = pd.read_csv(pd.io.common.StringIO(r.text))
        result = []
        for _, row in df.iterrows():
            ticker = str(row.get("Symbol", row.get("Ticker", ""))).strip()
            if not ticker or "." not in ticker:
                continue
            result.append({
                "ticker":   ticker,
                "name":     str(row.get("Name", row.get("Company", ""))).strip(),
                "country":  str(row.get("Country", "")).strip(),
                "sector":   "—",
                "exchange": EXCHANGE_NAMES.get(_ticker_suffix(ticker), ""),
                "mcap_b":   None,
                "cap_band": "",
            })
        print(f"  Live fetch: {len(result)} EU tickers")
        return result
    except Exception as e:
        print(f"  ⚠️  Live fetch failed: {e}")
        return []


def build_universe(universe_file=None, min_cap_b=0.0, exchange_filter=None) -> list[dict]:
    cached = _load_cache()
    if cached and not exchange_filter and min_cap_b == 0:
        print(f"  Symbol cache hit: {len(cached)} EU tickers")
        return cached

    path = universe_file or _DEFAULT_UNIVERSE_FILE
    if Path(path).exists():
        print(f"  Loading universe from {Path(path).name} …")
        try:
            universe = load_universe_from_file(Path(path))
        except Exception as e:
            print(f"  ⚠️  File read error: {e}")
            universe = []
    else:
        universe = []

    if not universe:
        universe = fetch_universe_live()

    if not universe:
        sys.exit("❌  Could not load EU universe. Supply the Excel file or check network.")

    print(f"  Raw universe: {len(universe)} tickers")

    # Deduplicate
    seen, deduped = set(), []
    for s in universe:
        if s["ticker"] not in seen:
            seen.add(s["ticker"])
            deduped.append(s)
    universe = deduped

    if min_cap_b > 0:
        before = len(universe)
        universe = [s for s in universe if s.get("mcap_b") is None or (s["mcap_b"] or 0) >= min_cap_b]
        print(f"  After mcap >= ${min_cap_b}B filter: {len(universe)} (dropped {before - len(universe)})")

    if exchange_filter:
        suffixes = {"." + e.lstrip(".").upper() for e in exchange_filter}
        before = len(universe)
        universe = [s for s in universe
                    if any(s["ticker"].upper().endswith(sfx) for sfx in suffixes)]
        print(f"  After exchange filter {suffixes}: {len(universe)} (dropped {before - len(universe)})")

    if not exchange_filter and min_cap_b == 0:
        _save_cache(universe)

    print(f"  → {len(universe)} EU tickers to scan")
    return universe


# ── Bulk OHLC ─────────────────────────────────────────────────────────────────

def bulk_download_ohlc(tickers: list, period: str = "3mo") -> dict:
    result = {}
    batches = [tickers[i:i + BATCH_SIZE] for i in range(0, len(tickers), BATCH_SIZE)]
    print(f"  Downloading OHLC for {len(tickers)} tickers in {len(batches)} batches …")

    for idx, batch in enumerate(batches, 1):
        print(f"    Batch {idx}/{len(batches)} ({len(batch)}) …", end=" ", flush=True)
        try:
            raw = yf.download(batch, period=period, auto_adjust=True,
                              threads=True, progress=False)
            if raw.empty:
                print("empty")
                continue
            if isinstance(raw.columns, pd.MultiIndex):
                for tkr in batch:
                    try:
                        df = raw.xs(tkr, axis=1, level=1).dropna(how="all")
                        if not df.empty and len(df) >= DARVAS_CONFIRM + 5:
                            result[tkr] = df
                    except KeyError:
                        pass
            else:
                tkr = batch[0]
                if not raw.empty:
                    result[tkr] = raw
            print(f"OK ({sum(1 for t in batch if t in result)} usable)")
        except Exception as e:
            print(f"ERROR — {e}")
        if idx < len(batches):
            time.sleep(SLEEP_BETWEEN)

    return result


# ── Darvas Box ────────────────────────────────────────────────────────────────

def compute_darvas_box(df, confirm=DARVAS_CONFIRM):
    if df is None or df.empty or len(df) < confirm + 5:
        return {"signal": "INSUFFICIENT_DATA", "box_top": None, "box_bottom": None}

    highs  = pd.to_numeric(df["High"],  errors="coerce").fillna(0).tolist()
    lows   = pd.to_numeric(df["Low"],   errors="coerce").fillna(0).tolist()
    closes = pd.to_numeric(df["Close"], errors="coerce").fillna(0).tolist()

    current = closes[-1]
    h, l = highs[:-1], lows[:-1]   # exclude current bar
    n = len(h)

    box_top_idx = box_top = None
    for i in range(n - confirm - 1, -1, -1):
        c = h[i]
        if c == 0:
            continue
        w = h[i + 1: i + 1 + confirm]
        if len(w) == confirm and all(x < c for x in w):
            box_top_idx, box_top = i, c
            break

    if box_top is None:
        return {"signal": "NO_BOX", "box_top": None, "box_bottom": None,
                "current_price": current}

    seg = l[box_top_idx:]
    box_bottom = None
    for i in range(len(seg) - confirm):
        c = seg[i]
        if c == 0:
            continue
        w = seg[i + 1: i + 1 + confirm]
        if len(w) == confirm and all(x > c for x in w):
            box_bottom = c
            break
    if box_bottom is None:
        valid = [x for x in seg if x > 0]
        box_bottom = min(valid) if valid else None

    if box_bottom is None:
        return {"signal": "NO_BOX", "box_top": round(box_top, 2), "box_bottom": None,
                "current_price": round(current, 2)}

    signal = ("BREAKOUT_BUY"   if current > box_top   else
              "BREAKDOWN_SELL" if current < box_bottom else "IN_BOX")
    rng    = box_top - box_bottom
    upside = ((box_top - current) / current * 100) if current else 0
    pos    = ((current - box_bottom) / rng * 100)  if rng    else 0

    return {
        "signal":       signal,
        "box_top":      round(box_top,    2),
        "box_bottom":   round(box_bottom, 2),
        "current_price":round(current,    2),
        "upside_pct":   round(upside, 2),
        "pos_in_box":   round(pos,    1),
        "data_points":  len(closes),
    }


# ── Fundamental helpers ───────────────────────────────────────────────────────

def _first_df(ticker, *attrs):
    for attr in attrs:
        df = getattr(ticker, attr, None)
        if df is not None and isinstance(df, pd.DataFrame) and not df.empty:
            return df
    return None


def _row(df, *names, col=0):
    if df is None or df.empty:
        return None
    for name in names:
        if name in df.index:
            try:
                val = df.loc[name].iloc[col]
                return float(val) if pd.notna(val) else None
            except Exception:
                pass
    return None


def _series(df, *names):
    if df is None or df.empty:
        return []
    for name in names:
        if name in df.index:
            return [float(v) for v in df.loc[name].dropna() if pd.notna(v)]
    return []


# ── Fundamental scan ─────────────────────────────────────────────────────────

def fundamental_scan(ticker_str: str) -> dict:
    result = {"ticker": ticker_str, "f_score": None, "cc_qualifies": "FAIL", "error": None}
    try:
        t = yf.Ticker(ticker_str)
        inc = _first_df(t, "income_stmt", "financials")
        bal = _first_df(t, "balance_sheet")
        cf  = _first_df(t, "cash_flow", "cashflow")
        if inc is None:
            result["error"] = "no_financial_data"
            return result
    except Exception as e:
        result["error"] = str(e)[:80]
        return result

    # ── Piotroski F-Score ─────────────────────────────────────────────────────
    try:
        ni0 = _row(inc, "Net Income", col=0);  a0 = _row(bal, "Total Assets", col=0)
        ni1 = _row(inc, "Net Income", col=1);  a1 = _row(bal, "Total Assets", col=1)
        roa0 = (ni0 / a0) if (ni0 and a0) else None
        roa1 = (ni1 / a1) if (ni1 and a1) else None
        ocf0 = _row(cf, "Operating Cash Flow", "Total Cash From Operating Activities")

        f1 = 1 if (roa0 and roa0 > 0)             else 0
        f2 = 1 if (ocf0 and ocf0 > 0)             else 0
        f3 = 1 if (roa0 and roa1 and roa0 > roa1) else 0
        f4 = 1 if (ocf0 and a0 and roa0 and (ocf0 / a0) > roa0) else 0

        ltd0 = _row(bal, "Long Term Debt", col=0) or 0
        ltd1 = _row(bal, "Long Term Debt", col=1) or 0
        f5 = 1 if (a0 and a1 and (ltd0 / a0) < (ltd1 / a1)) else 0

        ca0 = _row(bal, "Current Assets", "Total Current Assets", col=0)
        cl0 = _row(bal, "Current Liabilities", "Total Current Liabilities", col=0)
        ca1 = _row(bal, "Current Assets", "Total Current Assets", col=1)
        cl1 = _row(bal, "Current Liabilities", "Total Current Liabilities", col=1)
        f6 = 1 if (ca0 and cl0 and ca1 and cl1 and (ca0 / cl0) > (ca1 / cl1)) else 0

        sh0 = _row(bal, "Share Issued", col=0)
        sh1 = _row(bal, "Share Issued", col=1)
        f7 = (1 if sh0 <= sh1 else 0) if (sh0 and sh1) else 1

        rev0 = _row(inc, "Total Revenue", col=0); gp0 = _row(inc, "Gross Profit", col=0)
        rev1 = _row(inc, "Total Revenue", col=1); gp1 = _row(inc, "Gross Profit", col=1)
        f8 = 1 if (gp0 and rev0 and gp1 and rev1 and (gp0 / rev0) > (gp1 / rev1)) else 0
        f9 = 1 if (rev0 and a0 and rev1 and a1 and (rev0 / a0) > (rev1 / a1)) else 0

        f_score = f1 + f2 + f3 + f4 + f5 + f6 + f7 + f8 + f9
        result["f_score"]  = f_score
        result["f_strong"] = f_score >= PIOTROSKI_STRONG
    except Exception:
        return result

    if not result.get("f_strong"):
        result["cc_score"] = "—"
        return result

    # ── Coffee Can (Europe) ───────────────────────────────────────────────────
    try:
        c = {}
        revs = _series(inc, "Total Revenue")
        if len(revs) >= 2 and revs[-1] > 0:
            cagr = ((revs[0] / revs[-1]) ** (1 / (len(revs) - 1)) - 1) * 100
            c["C1"] = 1 if cagr > 10 else 0
        else:
            cagr = None; c["C1"] = 0

        ebit_s = _series(inc, "EBIT", "Operating Income", "Ebit")
        ta_s   = _series(bal, "Total Assets")
        cl_s   = _series(bal, "Current Liabilities", "Total Current Liabilities")
        roce_l = [ebit_s[i] / (ta_s[i] - cl_s[i]) * 100
                  for i in range(min(len(ebit_s), len(ta_s), len(cl_s)))
                  if (ta_s[i] - cl_s[i]) > 0]
        avg_roce = sum(roce_l) / len(roce_l) if roce_l else None
        c["C2"] = 1 if (avg_roce and avg_roce > 15) else 0

        ltd_s = _series(bal, "Long Term Debt")
        eq_s  = _series(bal, "Stockholders Equity", "Total Stockholder Equity",
                        "Total Equity Gross Minority Interest")
        de = (ltd_s[0] / abs(eq_s[0])) if (ltd_s and eq_s and eq_s[0] != 0) else None
        c["C3"] = 1 if (de is not None and de < 1) else 0

        try:
            mcap = t.fast_info.market_cap or 0
        except Exception:
            mcap = 0
        c["C4"] = 1 if mcap >= 1e9 else 0   # >= €1B

        ni_s = _series(inc, "Net Income")
        c["C5"] = 1 if (ni_s and all(n > 0 for n in ni_s)) else 0

        fcf_s = _series(cf, "Free Cash Flow")
        if fcf_s:
            c["C6"] = 1 if fcf_s[0] > 0 else 0
        else:
            ocf_s   = _series(cf, "Operating Cash Flow")
            capex_s = _series(cf, "Capital Expenditure")
            c["C6"] = 1 if (ocf_s and capex_s and (ocf_s[0] - abs(capex_s[0])) > 0) else 0

        qualifies = sum(c.values()) == len(c)
        result.update({
            "cc_qualifies":   "PASS" if qualifies else "FAIL",
            "cc_score":       f"{sum(c.values())}/{len(c)}",
            "cc_rev_cagr":    round(cagr, 2) if cagr else None,
            "cc_roce_avg":    round(avg_roce, 2) if avg_roce else None,
            "cc_debt_equity": round(de, 2) if de is not None else None,
            "market_cap_b":   round(mcap / 1e9, 2) if mcap else None,
        })
    except Exception:
        pass

    return result


# ── Excel styling ─────────────────────────────────────────────────────────────

def style_sheet(ws):
    if not OPENPYXL_OK:
        return
    fill_hdr  = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_alt  = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    font_hdr  = Font(name="Calibri", size=11, bold=True,  color="FFFFFF")
    font_body = Font(name="Calibri", size=11, bold=False, color="000000")
    thin = Border(
        left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),  bottom=Side(style="thin", color="E2E8F0"),
    )
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_hdr; cell.fill = fill_hdr; cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_body; cell.border = thin
            if row_idx % 2 == 1:
                cell.fill = fill_alt
            hdr = str(ws.cell(row=1, column=col_idx).value or "").upper()
            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if any(k in hdr for k in ["%", "CAGR", "ROCE", "YIELD"]):
                    cell.number_format = '0.00"%"'
                elif any(k in hdr for k in ["CAP", "LTP", "BOX", "PRICE", "200", "MCAP"]):
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.00'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Full European equity scanner — Darvas + Piotroski + Coffee Can"
    )
    parser.add_argument("--top",          type=int,   default=0,   help="Limit to first N tickers")
    parser.add_argument("--min-cap",      type=float, default=0.0, help="Min market cap $B (e.g. 1)")
    parser.add_argument("--exchange",     nargs="+",  default=[],  help="Filter suffixes e.g. PA DE ST")
    parser.add_argument("--no-scans",     action="store_true",     help="Skip Piotroski + Coffee Can")
    parser.add_argument("--workers",      type=int,   default=MAX_WORKERS)
    parser.add_argument("--universe-file",default=None, help="Path to EU universe Excel file")
    args = parser.parse_args()

    print(f"\n{'#'*60}")
    print(f"  FULL EUROPE MARKET SCAN")
    print(f"  Started: {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"{'#'*60}\n")

    # Stage 1 — Universe
    print("Stage 1 — Building EU equity universe …")
    universe = build_universe(
        universe_file=args.universe_file,
        min_cap_b=args.min_cap,
        exchange_filter=args.exchange,
    )
    meta    = {s["ticker"]: s for s in universe}
    tickers = [s["ticker"] for s in universe]

    if args.top:
        tickers = tickers[:args.top]
        print(f"  (limited to first {args.top} tickers)")

    # Exchange breakdown summary
    suffix_counts = Counter(_ticker_suffix(t) for t in tickers)
    summary = "  ".join(
        f"{EXCHANGE_NAMES.get(sfx, sfx)} {n}"
        for sfx, n in suffix_counts.most_common(8)
    )
    print(f"  Exchanges: {summary}")

    # Stage 2 — Bulk OHLC
    print(f"\nStage 2 — Bulk OHLC download ({len(tickers)} tickers) …")
    ohlc = bulk_download_ohlc(tickers, period="3mo")
    print(f"  → {len(ohlc)} tickers with usable OHLC data")

    # Stage 3 — Darvas + MA
    print("\nStage 3 — Darvas Box & 200-day MA screen …")
    all_rows, darvas_rows, breakout_tickers = [], [], []

    for tkr, df in ohlc.items():
        info    = meta.get(tkr, {})
        darv    = compute_darvas_box(df)
        closes  = pd.to_numeric(df["Close"], errors="coerce").dropna()
        ltp     = round(float(closes.iloc[-1]), 2) if not closes.empty else None
        prev    = round(float(closes.iloc[-2]), 2) if len(closes) >= 2 else None
        chg_pct = round((ltp - prev) / prev * 100, 2) if (ltp and prev) else None

        ma200   = round(closes.rolling(200).mean().iloc[-1], 2) if len(closes) >= 200 else None
        dist_ma = round((ltp - ma200) / ma200 * 100, 2) if (ma200 and ltp) else None
        trend   = ("Above 200MA (Uptrend)"      if dist_ma and dist_ma >  5 else
                   "Below 200MA (Downtrend)"    if dist_ma and dist_ma < -5 else
                   "Near 200MA (Consolidation)" if dist_ma else "Insufficient History")

        sfx            = _ticker_suffix(tkr)
        exchange_label = EXCHANGE_NAMES.get(sfx, sfx)

        row = {
            "Ticker":             tkr,
            "Name":               info.get("name", ""),
            "Country":            info.get("country", ""),
            "Sector":             info.get("sector", "—"),
            "Exchange":           exchange_label,
            "Cap_Band":           info.get("cap_band", ""),
            "MCap_B_USD":         info.get("mcap_b"),
            "LTP":                ltp,
            "Change%":            chg_pct,
            "200_Day_MA":         ma200,
            "Distance_to_200MA%": dist_ma,
            "Trend_Signal":       trend,
            "Darvas_Signal":      darv.get("signal"),
            "Box_Top":            darv.get("box_top"),
            "Box_Bottom":         darv.get("box_bottom"),
            "Upside_to_Top%":     darv.get("upside_pct"),
            "Position_in_Box%":   darv.get("pos_in_box"),
            "Data_Points":        darv.get("data_points"),
        }
        all_rows.append(row)
        if darv.get("signal") in ("BREAKOUT_BUY", "BREAKDOWN_SELL"):
            darvas_rows.append(row.copy())
        if darv.get("signal") == "BREAKOUT_BUY":
            breakout_tickers.append(tkr)

    breakdowns = sum(1 for r in darvas_rows if r["Darvas_Signal"] == "BREAKDOWN_SELL")
    print(f"  Breakout BUY:   {len(breakout_tickers)}")
    print(f"  Breakdown SELL: {breakdowns}")
    print(f"  In Box:         {len(all_rows) - len(darvas_rows)}")

    # Stage 4 — Fundamentals
    fund_rows, triple_rows = [], []

    if not args.no_scans and breakout_tickers:
        def _upside(tkr):
            r = next((x for x in darvas_rows if x["Ticker"] == tkr), None)
            return abs(r.get("Upside_to_Top%") or 999)

        if len(breakout_tickers) > MAX_FUND_CANDIDATES:
            breakout_tickers = sorted(breakout_tickers, key=_upside)[:MAX_FUND_CANDIDATES]
            print(f"  (capped to {MAX_FUND_CANDIDATES} freshest breakouts for fundamentals)")

        print(f"\nStage 4 — Fundamental scans ({len(breakout_tickers)} candidates, "
              f"{args.workers} workers) …")
        done = 0
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            futures = {pool.submit(fundamental_scan, tkr): tkr for tkr in breakout_tickers}
            for future in as_completed(futures):
                tkr  = futures[future]
                done += 1
                try:
                    res  = future.result()
                    tech = next((r for r in darvas_rows if r["Ticker"] == tkr), {})
                    fund_row = {
                        "Ticker":          tkr,
                        "Name":            tech.get("Name"),
                        "Country":         tech.get("Country"),
                        "Sector":          tech.get("Sector"),
                        "Exchange":        tech.get("Exchange"),
                        "Cap_Band":        tech.get("Cap_Band"),
                        "MCap_B_USD":      tech.get("MCap_B_USD"),
                        "LTP":             tech.get("LTP"),
                        "Change%":         tech.get("Change%"),
                        "Darvas_Signal":   tech.get("Darvas_Signal"),
                        "Upside_to_Top%":  tech.get("Upside_to_Top%"),
                        "Box_Top":         tech.get("Box_Top"),
                        "200_Day_MA":      tech.get("200_Day_MA"),
                        "Piotroski_Score": res.get("f_score"),
                        "Piotroski_Strong":("YES" if res.get("f_strong") else "NO"),
                        "CoffeeCan":       res.get("cc_qualifies", "FAIL"),
                        "CC_Score":        res.get("cc_score", "—"),
                        "Rev_CAGR%":       res.get("cc_rev_cagr"),
                        "ROCE_Avg%":       res.get("cc_roce_avg"),
                        "Debt_Equity":     res.get("cc_debt_equity"),
                        "MCap_B_yf":       res.get("market_cap_b"),
                        "Error":           res.get("error", ""),
                    }
                    fund_rows.append(fund_row)
                    if res.get("f_strong") and res.get("cc_qualifies") == "PASS":
                        triple_rows.append(fund_row.copy())
                    if done % 25 == 0 or done == len(breakout_tickers):
                        print(f"    {done}/{len(breakout_tickers)} done  "
                              f"(triple hits: {len(triple_rows)})")
                except Exception as e:
                    print(f"    {tkr}: error — {e}")
    else:
        print("\nStage 4 — Skipped")

    # Stage 5 — Excel
    print("\nStage 5 — Saving Excel workbook …")
    date_str = datetime.today().strftime("%Y%m%d_%H%M")
    path = DOWNLOAD_DIR / f"europe_market_scan_{date_str}.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        def write_sheet(rows, name, sort_col=None):
            if not rows:
                pd.DataFrame().to_excel(writer, sheet_name=name, index=False)
                return
            df = pd.DataFrame(rows)
            if sort_col and sort_col in df.columns:
                df = df.sort_values(sort_col, ascending=False)
            df.to_excel(writer, sheet_name=name, index=False)
            if OPENPYXL_OK:
                style_sheet(writer.sheets[name])

        write_sheet(all_rows,    "All_Stocks",    sort_col="Change%")
        write_sheet(darvas_rows, "Darvas_Signals", sort_col="Upside_to_Top%")
        write_sheet(fund_rows,   "Fundamentals",  sort_col="Piotroski_Score")
        write_sheet(triple_rows, "Triple_Hits",   sort_col="Piotroski_Score")

    print(f"\n{'='*60}")
    print(f"  SCAN COMPLETE — {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"  EU stocks scanned:     {len(all_rows)}")
    print(f"  Darvas Breakouts:      {len(breakout_tickers)}")
    print(f"  Fundamentals scanned:  {len(fund_rows)}")
    print(f"  ★ TRIPLE HITS:         {len(triple_rows)}")
    if triple_rows:
        print("\n  Triple-hit stocks:")
        for r in sorted(triple_rows, key=lambda x: x.get("Piotroski_Score") or 0, reverse=True):
            print(f"    {r['Ticker']:<14} {r['Name']:<30}  "
                  f"F={r['Piotroski_Score']}/9  CC={r['CC_Score']}  "
                  f"{r['Country']}  {r['Exchange']}")
    else:
        print("\n  No Triple Hit stocks found today.")
    print(f"\n  📊 Excel saved → {path}\n")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
